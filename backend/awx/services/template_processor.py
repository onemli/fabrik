# awx/services/template_processor.py
#
# Pre-execution processing for AutomationTemplate schemas: validates that the
# schema structure is sane, transforms user-submitted data into the shape the
# execution engine expects, and parses Excel uploads into the same dict format
# as the wizard UI produces.
#
# Excel parsing uses openpyxl because it's the only library that handles modern
# .xlsx without a C extension. Rows are matched to columns by header name, not
# position, so reordering columns in the spreadsheet doesn't break the import.

import logging
import re
from typing import Dict, List, Tuple, Any
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

logger = logging.getLogger(__name__)


class TemplateProcessorError(Exception):
    pass


class SchemaValidationError(TemplateProcessorError):
    pass


class DataTransformationError(TemplateProcessorError):
    pass


class ExcelParsingError(TemplateProcessorError):
    pass


class TemplateProcessor:
    def __init__(self):
        self.logger = logging.getLogger(__name__)

    def validate_table_schema(self, schema: List[Dict[str, Any]]) -> Tuple[bool, List[str]]:
        errors = []

        try:
            # Check if schema is a list
            if not isinstance(schema, list):
                raise SchemaValidationError('Schema must be a list of table definitions')

            # Check if schema is empty
            if len(schema) == 0:
                errors.append('Schema cannot be empty')
                return False, errors

            # Validate each table
            for idx, table in enumerate(schema):
                table_errors = self._validate_table_definition(table, idx)
                errors.extend(table_errors)

            is_valid = len(errors) == 0

            if is_valid:
                self.logger.info(f'Schema validation successful: {len(schema)} tables validated')
            else:
                self.logger.warning(f'Schema validation failed with {len(errors)} errors')

            return is_valid, errors

        except Exception as e:
            self.logger.error(f'Schema validation error: {str(e)}', exc_info=True)
            raise SchemaValidationError(f'Critical schema validation error: {str(e)}') from e

    def _validate_table_definition(self, table: Dict[str, Any], idx: int) -> List[str]:
        errors = []

        try:
            # Check required table fields
            if not isinstance(table, dict):
                errors.append(f'Table {idx}: Must be a dictionary')
                return errors

            # Validate table_name
            if 'table_name' not in table:
                errors.append(f"Table {idx}: Missing 'table_name' field")
            elif not isinstance(table['table_name'], str):
                errors.append(f"Table {idx}: 'table_name' must be a string")
            elif not table['table_name'].strip():
                errors.append(f"Table {idx}: 'table_name' cannot be empty")
            elif not re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', table['table_name']):
                errors.append(
                    f"Table {idx}: 'table_name' must be a valid identifier (alphanumeric and underscore)"
                )

            # Validate columns
            if 'columns' not in table:
                errors.append(f"Table {idx}: Missing 'columns' field")
            elif not isinstance(table['columns'], list):
                errors.append(f"Table {idx}: 'columns' must be a list")
            elif len(table['columns']) == 0:
                errors.append(f"Table {idx}: 'columns' cannot be empty")
            else:
                # Validate each column
                column_names = set()
                for col_idx, column in enumerate(table['columns']):
                    col_errors = self._validate_column_definition(
                        column, idx, col_idx, column_names
                    )
                    errors.extend(col_errors)

            # Validate optional fields
            if 'min_rows' in table:
                if not isinstance(table['min_rows'], int) or table['min_rows'] < 0:
                    errors.append(f"Table {idx}: 'min_rows' must be a non-negative integer")

            if 'max_rows' in table:
                if not isinstance(table['max_rows'], int) or table['max_rows'] < 1:
                    errors.append(f"Table {idx}: 'max_rows' must be a positive integer")

                # Check min_rows <= max_rows
                if 'min_rows' in table and table['min_rows'] > table['max_rows']:
                    errors.append(f"Table {idx}: 'min_rows' cannot be greater than 'max_rows'")

            return errors

        except Exception as e:
            self.logger.error(f'Error validating table {idx}: {str(e)}', exc_info=True)
            return [f'Table {idx}: Unexpected validation error: {str(e)}']

    def _validate_column_definition(
        self, column: Dict[str, Any], table_idx: int, col_idx: int, column_names: set
    ) -> List[str]:
        errors = []

        try:
            # Check if column is a dict
            if not isinstance(column, dict):
                errors.append(f'Table {table_idx}, Column {col_idx}: Must be a dictionary')
                return errors

            # Validate name
            if 'name' not in column:
                errors.append(f"Table {table_idx}, Column {col_idx}: Missing 'name' field")
            elif not isinstance(column['name'], str):
                errors.append(f"Table {table_idx}, Column {col_idx}: 'name' must be a string")
            elif not column['name'].strip():
                errors.append(f"Table {table_idx}, Column {col_idx}: 'name' cannot be empty")
            elif not re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', column['name']):
                errors.append(
                    f"Table {table_idx}, Column {col_idx}: 'name' must be a valid identifier"
                )
            else:
                # Check for duplicate column names
                if column['name'] in column_names:
                    errors.append(
                        f"Table {table_idx}, Column {col_idx}: Duplicate column name '{column['name']}'"
                    )
                else:
                    column_names.add(column['name'])

            # Validate type
            valid_types = ['text', 'number', 'boolean', 'enum']
            if 'type' not in column:
                errors.append(f"Table {table_idx}, Column {col_idx}: Missing 'type' field")
            elif column['type'] not in valid_types:
                errors.append(
                    f"Table {table_idx}, Column {col_idx}: 'type' must be one of {valid_types}"
                )

            # Validate required
            if 'required' in column and not isinstance(column['required'], bool):
                errors.append(f"Table {table_idx}, Column {col_idx}: 'required' must be boolean")

            # Validate validation regex
            if 'validation' in column:
                if not isinstance(column['validation'], str):
                    errors.append(
                        f"Table {table_idx}, Column {col_idx}: 'validation' must be a string (regex)"
                    )
                else:
                    # Test if regex is valid
                    try:
                        re.compile(column['validation'])
                    except re.error as e:
                        errors.append(
                            f'Table {table_idx}, Column {col_idx}: Invalid regex pattern: {str(e)}'
                        )

            # Validate enum values
            if column.get('type') == 'enum':
                if 'enum_values' not in column:
                    errors.append(
                        f"Table {table_idx}, Column {col_idx}: 'enum' type requires 'enum_values' field"
                    )
                elif not isinstance(column['enum_values'], list):
                    errors.append(
                        f"Table {table_idx}, Column {col_idx}: 'enum_values' must be a list"
                    )
                elif len(column['enum_values']) == 0:
                    errors.append(
                        f"Table {table_idx}, Column {col_idx}: 'enum_values' cannot be empty"
                    )

            return errors

        except Exception as e:
            self.logger.error(f'Error validating column: {str(e)}', exc_info=True)
            return [f'Table {table_idx}, Column {col_idx}: Unexpected validation error: {str(e)}']

    def transform_input_to_ansible_vars(
        self, input_data: Dict[str, List[Dict[str, Any]]], mappings: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        Transform user input data to Ansible extra_vars format.

        Args:
            input_data: User input data organized by table name
            mappings: Variable mappings configuration

        Returns:
            Ansible extra_vars dictionary

        Raises:
            DataTransformationError: If transformation fails
        """
        try:
            extra_vars = {}

            # Validate inputs
            if not isinstance(input_data, dict):
                raise DataTransformationError('input_data must be a dictionary')

            if not isinstance(mappings, dict):
                raise DataTransformationError('mappings must be a dictionary')

            # Transform each table
            for table_name, table_data in input_data.items():
                if not isinstance(table_data, list):
                    raise DataTransformationError(f"Table '{table_name}' data must be a list")

                # Get mapping for this table
                if table_name not in mappings:
                    self.logger.warning(f"No mapping found for table '{table_name}', skipping")
                    continue

                table_mapping = mappings[table_name]

                # Validate mapping structure
                if not isinstance(table_mapping, dict):
                    raise DataTransformationError(
                        f"Mapping for table '{table_name}' must be a dictionary"
                    )

                if 'target_var' not in table_mapping:
                    raise DataTransformationError(
                        f"Mapping for table '{table_name}' missing 'target_var'"
                    )

                if 'column_mapping' not in table_mapping:
                    raise DataTransformationError(
                        f"Mapping for table '{table_name}' missing 'column_mapping'"
                    )

                target_var = table_mapping['target_var']
                column_mapping = table_mapping['column_mapping']

                # Transform rows
                transformed_rows = []
                for row_idx, row in enumerate(table_data):
                    try:
                        transformed_row = self._transform_row(row, column_mapping)
                        transformed_rows.append(transformed_row)
                    except Exception as e:
                        raise DataTransformationError(
                            f"Error transforming row {row_idx} in table '{table_name}': {str(e)}"
                        ) from e

                # Add to extra_vars
                extra_vars[target_var] = transformed_rows

            self.logger.info(f'Successfully transformed {len(input_data)} tables to Ansible vars')
            return extra_vars

        except DataTransformationError:
            raise
        except Exception as e:
            self.logger.error(f'Data transformation error: {str(e)}', exc_info=True)
            raise DataTransformationError(f'Unexpected transformation error: {str(e)}') from e

    def _transform_row(self, row: Dict[str, Any], column_mapping: Dict[str, str]) -> Dict[str, Any]:
        """Transform a single row using column mapping."""
        transformed = {}

        for input_col, ansible_var in column_mapping.items():
            if input_col in row:
                value = row[input_col]
                if value is not None:
                    transformed[ansible_var] = value

        return transformed

    def parse_excel_file(
        self, file_bytes: bytes, schema: List[Dict[str, Any]]
    ) -> Tuple[Dict[str, List[Dict[str, Any]]], List[str]]:
        """
        Parse Excel file and validate against schema.

        Args:
            file_bytes: Excel file content as bytes
            schema: Template table schema

        Returns:
            Tuple of (parsed_data, error_messages)

        Raises:
            ExcelParsingError: If file cannot be parsed
        """
        errors = []
        parsed_data = {}

        try:
            # Load workbook
            try:
                workbook = load_workbook(BytesIO(file_bytes), data_only=True)
            except InvalidFileException as e:
                raise ExcelParsingError(f'Invalid Excel file: {str(e)}') from e
            except Exception as e:
                raise ExcelParsingError(f'Failed to open Excel file: {str(e)}') from e

            # Parse each table (sheet)
            for table_def in schema:
                table_name = table_def['table_name']

                # Check if sheet exists
                if table_name not in workbook.sheetnames:
                    errors.append(f"Sheet '{table_name}' not found in Excel file")
                    continue

                try:
                    sheet = workbook[table_name]
                    table_data, table_errors = self._parse_sheet(sheet, table_def)

                    parsed_data[table_name] = table_data
                    errors.extend(table_errors)

                except Exception as e:
                    errors.append(f"Error parsing sheet '{table_name}': {str(e)}")
                    self.logger.error(f'Sheet parsing error: {str(e)}', exc_info=True)

            if len(errors) == 0:
                self.logger.info(f'Successfully parsed Excel file with {len(parsed_data)} tables')
            else:
                self.logger.warning(f'Excel parsing completed with {len(errors)} errors')

            return parsed_data, errors

        except ExcelParsingError:
            raise
        except Exception as e:
            self.logger.error(f'Excel parsing error: {str(e)}', exc_info=True)
            raise ExcelParsingError(f'Unexpected Excel parsing error: {str(e)}') from e

    def _parse_sheet(
        self, sheet, table_def: Dict[str, Any]
    ) -> Tuple[List[Dict[str, Any]], List[str]]:
        """Parse a single Excel sheet."""
        errors = []
        rows_data = []

        try:
            columns = table_def['columns']

            # Read header row (row 1)
            header_row = [cell.value for cell in sheet[1]]

            # Create column mapping
            column_indices = {}
            for col_def in columns:
                col_name = col_def['name']
                try:
                    col_index = header_row.index(col_name)
                    column_indices[col_name] = col_index
                except ValueError:
                    if col_def.get('required', False):
                        errors.append(f"Required column '{col_name}' not found in sheet")

            # Read data rows (starting from row 2)
            for row_idx in range(2, sheet.max_row + 1):
                row = sheet[row_idx]
                row_data = {}
                row_empty = True

                # Read each column
                for col_def in columns:
                    col_name = col_def['name']

                    if col_name not in column_indices:
                        continue

                    col_index = column_indices[col_name]
                    cell_value = row[col_index].value

                    # Check if row has any data
                    if cell_value is not None and str(cell_value).strip():
                        row_empty = False

                    # Validate and convert value
                    try:
                        converted_value = self._convert_cell_value(
                            cell_value, col_def, row_idx, col_name
                        )
                        row_data[col_name] = converted_value
                    except ValueError as e:
                        errors.append(f"Row {row_idx}, Column '{col_name}': {str(e)}")

                # Skip empty rows
                if not row_empty:
                    rows_data.append(row_data)

            # Validate row count
            min_rows = table_def.get('min_rows', 1)
            max_rows = table_def.get('max_rows', float('inf'))

            if len(rows_data) < min_rows:
                errors.append(f'Table requires at least {min_rows} rows, got {len(rows_data)}')

            if len(rows_data) > max_rows:
                errors.append(f'Table allows maximum {max_rows} rows, got {len(rows_data)}')

            return rows_data, errors

        except Exception as e:
            self.logger.error(f'Sheet parsing error: {str(e)}', exc_info=True)
            return [], [f'Error parsing sheet: {str(e)}']

    def _convert_cell_value(
        self, value: Any, col_def: Dict[str, Any], row_idx: int, col_name: str
    ) -> Any:
        """Convert and validate Excel cell value."""
        col_type = col_def['type']
        is_required = col_def.get('required', False)

        # Handle None/empty values
        if value is None or (isinstance(value, str) and not value.strip()):
            if is_required:
                raise ValueError('Required field cannot be empty')

            # Use default if provided
            if 'default' in col_def:
                return col_def['default']

            return None

        # Convert based on type
        try:
            if col_type == 'text':
                converted = str(value).strip()

                # Regex validation — skip if validation_mode is 'none'; prefer
                # the live RegexPattern record over the inline string so edits
                # to a shared pattern propagate without re-saving templates.
                validation_mode = col_def.get('validation_mode', 'regex')
                pattern = None
                if validation_mode == 'regex':
                    pattern_id = col_def.get('regex_pattern_id')
                    if pattern_id:
                        from awx.models.validation import RegexPattern

                        try:
                            pattern = (
                                RegexPattern.objects.only('pattern').get(pk=pattern_id).pattern
                            )
                        except RegexPattern.DoesNotExist:
                            pattern = None
                    if not pattern:
                        pattern = col_def.get('validation')
                if pattern and not re.match(pattern, converted):
                    raise ValueError(f"Value '{converted}' does not match pattern '{pattern}'")

                return converted

            elif col_type == 'number':
                if isinstance(value, (int, float)):
                    return value

                try:
                    if isinstance(value, str) and '.' not in value:
                        return int(value)
                    return float(value)
                except (ValueError, TypeError):
                    raise ValueError(f"Value '{value}' is not a valid number")

            elif col_type == 'boolean':
                if isinstance(value, bool):
                    return value

                str_value = str(value).lower().strip()
                if str_value in ['true', '1', 'yes', 'y']:
                    return True
                elif str_value in ['false', '0', 'no', 'n']:
                    return False
                else:
                    raise ValueError(f"Value '{value}' is not a valid boolean")

            elif col_type == 'enum':
                str_value = str(value).strip()
                enum_values = col_def.get('enum_values', [])

                if str_value not in enum_values:
                    raise ValueError(f"Value '{str_value}' must be one of {enum_values}")

                return str_value

            else:
                return value

        except ValueError:
            raise
        except Exception as e:
            raise ValueError(f'Conversion error: {str(e)}') from e

    def generate_excel_template(
        self, schema: List[Dict[str, Any]], template_name: str = 'Template'
    ) -> bytes:
        """Generate Excel template file from schema."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.worksheet.datavalidation import DataValidation

            workbook = Workbook()

            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']

            # Create a sheet for each table
            for table_def in schema:
                table_name = table_def['table_name']
                columns = table_def['columns']

                sheet = workbook.create_sheet(title=table_name)

                # Style definitions
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(
                    start_color='366092', end_color='366092', fill_type='solid'
                )
                header_alignment = Alignment(horizontal='center', vertical='center')

                border_side = Side(style='thin', color='000000')
                border = Border(
                    left=border_side, right=border_side, top=border_side, bottom=border_side
                )

                # Write headers
                for col_idx, col_def in enumerate(columns, start=1):
                    cell = sheet.cell(row=1, column=col_idx)

                    col_name = col_def['name']
                    if col_def.get('required', False):
                        col_name += ' *'

                    cell.value = col_name
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border

                    # Add column comment/tooltip
                    if 'description' in col_def:
                        from openpyxl.comments import Comment

                        cell.comment = Comment(col_def['description'], 'Template')

                    # Set column width
                    column_letter = sheet.cell(row=1, column=col_idx).column_letter
                    sheet.column_dimensions[column_letter].width = 20

                # Add example data if provided
                if 'example_data' in table_def:
                    for row_idx, example_row in enumerate(table_def['example_data'], start=2):
                        for col_idx, col_def in enumerate(columns, start=1):
                            col_name = col_def['name']
                            if col_name in example_row:
                                cell = sheet.cell(row=row_idx, column=col_idx)
                                cell.value = example_row[col_name]
                                cell.border = border

                # Add data validation for enum columns
                for col_idx, col_def in enumerate(columns, start=1):
                    if col_def['type'] == 'enum' and 'enum_values' in col_def:
                        column_letter = sheet.cell(row=1, column=col_idx).column_letter

                        dv = DataValidation(
                            type='list',
                            formula1=f'"{",".join(col_def["enum_values"])}"',
                            allow_blank=not col_def.get('required', False),
                        )
                        dv.error = 'Invalid value'
                        dv.errorTitle = 'Invalid Entry'
                        dv.prompt = f'Select from: {", ".join(col_def["enum_values"])}'
                        dv.promptTitle = 'Valid Values'

                        dv.add(f'{column_letter}2:{column_letter}1000')
                        sheet.add_data_validation(dv)

                # Freeze header row
                sheet.freeze_panes = 'A2'

            # Save to bytes
            output = BytesIO()
            workbook.save(output)
            excel_bytes = output.getvalue()

            self.logger.info(f'Generated Excel template with {len(schema)} sheets')
            return excel_bytes

        except Exception as e:
            self.logger.error(f'Excel template generation error: {str(e)}', exc_info=True)
            raise ExcelParsingError(f'Failed to generate Excel template: {str(e)}') from e
